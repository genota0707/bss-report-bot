import os
import json
import uuid
import threading
import requests
import openpyxl
from io import BytesIO
from flask import Flask, request, abort, send_file
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from linebot.v3 import WebhookHandler
from linebot.v3.exceptions import InvalidSignatureError
from linebot.v3.messaging import (
    Configuration,
    ApiClient,
    MessagingApi,
    ReplyMessageRequest,
    TextMessage
)
from linebot.v3.webhooks import MessageEvent, TextMessageContent

app = Flask(__name__)

LINE_CHANNEL_SECRET = os.getenv('LINE_CHANNEL_SECRET')
LINE_CHANNEL_ACCESS_TOKEN = os.getenv('LINE_CHANNEL_ACCESS_TOKEN')
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
GOOGLE_DRIVE_FILE_ID = os.getenv('GOOGLE_DRIVE_FILE_ID')
GOOGLE_SERVICE_ACCOUNT_JSON = os.getenv('GOOGLE_SERVICE_ACCOUNT_JSON')
SERVER_BASE_URL = os.getenv('RENDER_EXTERNAL_URL', 'https://bss-report-bot.onrender.com')

configuration = Configuration(access_token=LINE_CHANNEL_ACCESS_TOKEN)
handler = WebhookHandler(LINE_CHANNEL_SECRET)

TEMPLATE_PATH = "BASICサッカースクール_レポートフォーマット_A4ぴったり.xlsx"

@app.route("/", methods=['GET'])
def index():
    return "B.S.S Report Bot is running!"

@app.route("/files/<filename>", methods=['GET'])
def download_file(filename):
    file_path = os.path.join("/tmp", filename)
    if os.path.exists(file_path):
        return send_file(
            file_path,
            as_attachment=True,
            download_name="成長レポート.xlsx"
        )
    return "File not found", 404

@app.route("/callback", methods=['POST'])
def callback():
    signature = request.headers.get('X-Line-Signature')
    body = request.get_data(as_text=True)

    try:
        handler.handle(body, signature)
    except InvalidSignatureError:
        abort(400)

    return 'OK'

def fetch_memo_from_drive():
    if not GOOGLE_DRIVE_FILE_ID:
        return None, "環境変数 GOOGLE_DRIVE_FILE_ID が未設定です。"

    file_id = GOOGLE_DRIVE_FILE_ID.strip()

    # 1. サービスアカウント認証
    if GOOGLE_SERVICE_ACCOUNT_JSON:
        try:
            info = json.loads(GOOGLE_SERVICE_ACCOUNT_JSON)
            if "private_key" in info and isinstance(info["private_key"], str):
                info["private_key"] = info["private_key"].replace("\\n", "\n")

            creds = Credentials.from_service_account_info(
                info, scopes=['https://www.googleapis.com/auth/drive.readonly']
            )
            service = build('drive', 'v3', credentials=creds)

            try:
                request_file = service.files().export_media(
                    fileId=file_id,
                    mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                file_stream = BytesIO(request_file.execute())
                wb = openpyxl.load_workbook(file_stream, data_only=True)
                return wb, None
            except Exception as e_export:
                try:
                    request_file = service.files().get_media(fileId=file_id)
                    file_stream = BytesIO(request_file.execute())
                    wb = openpyxl.load_workbook(file_stream, data_only=True)
                    return wb, None
                except Exception as e_get:
                    return None, f"Drive取得エラー: export({str(e_export)}) / get({str(e_get)})"
        except Exception as e:
            return None, f"認証キー解析エラー: {str(e)}"

    # 2. 公開リンク取得
    pub_sheet_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    try:
        res = requests.get(pub_sheet_url, timeout=10)
        if res.status_code == 200 and len(res.content) > 500:
            wb = openpyxl.load_workbook(BytesIO(res.content), data_only=True)
            return wb, None
    except Exception:
        pass

    return None, "Googleドライブのファイル取得に失敗しました。"

def find_student_data(wb, name_query, month_query):
    target_sheet = None
    for sheet_name in wb.sheetnames:
        if month_query in sheet_name or f"{month_query}月" in sheet_name:
            target_sheet = wb[sheet_name]
            break
            
    if target_sheet is None:
        target_sheet = wb[wb.sheetnames[-1]]
    sheet = target_sheet

    extracted_info = {
        "name": name_query,
        "grade": "",
        "course": "",
        "month": month_query if "月" in month_query else f"{month_query}月",
        "score_stop": "",
        "score_kick": "",
        "score_carry": "",
        "score_judge": "",
        "memo": ""
    }

    clean_name_query = name_query.replace(" ", "").replace(" ", "")

    for r in range(1, 50):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 15)]
        row_str = "".join([str(v) for v in row_vals if v is not None])
        clean_row_str = row_str.replace(" ", "").replace(" ", "")
        
        if clean_name_query in clean_row_str:
            for val in row_vals:
                if val is not None:
                    s_val = str(val).strip()
                    if "年" in s_val and not extracted_info["grade"]:
                        extracted_info["grade"] = s_val
                    elif "週" in s_val and not extracted_info["course"]:
                        extracted_info["course"] = s_val
            
            scores = []
            memos = []
            for cell_val in row_vals[3:]:
                if isinstance(cell_val, (int, float)):
                    scores.append(int(cell_val))
                elif isinstance(cell_val, str) and cell_val.strip():
                    if not any(k in cell_val for k in ["年生", "週1回", "週2回"]):
                        memos.append(cell_val.strip())
            
            if len(scores) >= 4:
                extracted_info["score_stop"] = str(scores[0])
                extracted_info["score_kick"] = str(scores[1])
                extracted_info["score_carry"] = str(scores[2])
                extracted_info["score_judge"] = str(scores[3])
                
            if memos:
                extracted_info["memo"] = " ".join(memos)
            break

    return extracted_info

def generate_ai_report(student_data, api_key):
    prompt = f"""
以下のサッカースクールの生徒メモデータをもとに、保護者・生徒向けの成長レポート文章（今月の成長ポイント、来月の目標）を作成し、JSON形式で出力してください。

生徒データ:
- 名前: {student_data.get('name')}
- 学年: {student_data.get('grade')}
- 月: {student_data.get('month')}
- 評価 (止める:{student_data.get('score_stop')}, 蹴る:{student_data.get('score_kick')}, 運ぶ:{student_data.get('score_carry')}, 判断:{student_data.get('score_judge')})
- 今月の気付きメモ: {student_data.get('memo')}

出力フォーマット(JSONのみ出力してください。マークダウンなどの囲みは不要です):
{{
  "growth_point": "今月の成長ポイント(150文字程度の丁寧な保護者向け文章)",
  "next_goal": "来月の目標(150文字程度の丁寧な保護者向け文章)"
}}
"""
    payload = {"contents": [{"parts": [{"text": prompt}]}]}

    candidate_urls = []
    try:
        list_url = f"https://generativelanguage.googleapis.com/v1beta/models?key={api_key}"
        list_res = requests.get(list_url, timeout=5)
        if list_res.status_code == 200:
            models_list = list_res.json().get("models", [])
            for m in models_list:
                if "generateContent" in m.get("supportedGenerationMethods", []):
                    m_name = m.get("name")
                    candidate_urls.append(f"https://generativelanguage.googleapis.com/v1beta/{m_name}:generateContent?key={api_key}")
    except Exception:
        pass

    if not candidate_urls:
        candidate_urls = [f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}"]

    for url in candidate_urls:
        try:
            res = requests.post(url, json=payload, headers={"Content-Type": "application/json"}, timeout=10)
            res_data = res.json()
            if res.status_code == 200 and 'candidates' in res_data:
                raw_text = res_data['candidates'][0]['content']['parts'][0]['text'].strip()
                # 記号エスケープを安全に処理
                bq = "`" * 3
                raw_text = raw_text.replace(bq + "json", "").replace(bq, "").strip()
                return json.loads(raw_text)
        except Exception:
            continue

    return {"growth_point": "日々の練習で素晴らしい成長が見られます。", "next_goal": "基本技術のさらなる向上を目指します。"}

def create_excel_report(data):
    if not os.path.exists(TEMPLATE_PATH):
        return None, None

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    sheet = wb.active

    if data.get("name"): sheet["C5"] = data["name"]
    if data.get("grade"): sheet["H5"] = data["grade"]
    if data.get("month"): sheet["C6"] = data["month"]
    if data.get("course"): sheet["H6"] = data["course"]

    if data.get("score_stop"): sheet["E9"] = str(data["score_stop"])
    if data.get("score_kick"): sheet["E10"] = str(data["score_kick"])
    if data.get("score_carry"): sheet["E11"] = str(data["score_carry"])
    if data.get("score_judge"): sheet["E12"] = str(data["score_judge"])

    if data.get("growth_point"): sheet["B16"] = data["growth_point"]
    if data.get("next_goal"): sheet["B25"] = data["next_goal"]

    unique_id = str(uuid.uuid4())[:8]
    file_name = f"report_{unique_id}.xlsx"
    file_path = os.path.join("/tmp", file_name)
    wb.save(file_path)
    return file_name, file_path

def async_process_and_reply(reply_token, user_text):
    try:
        parts = user_text.strip().split()
        name_query = parts[0] if len(parts) > 0 else ""
        month_query = parts[1] if len(parts) > 1 else "8月"

        wb_memo, error_msg = fetch_memo_from_drive()
        
        if not wb_memo:
            reply_text = f"Googleドライブのメモ取得に失敗しました。\n【詳細】\n{error_msg}"
        else:
            student_data = find_student_data(wb_memo, name_query, month_query)
            ai_res = generate_ai_report(student_data, GEMINI_API_KEY)
            student_data["growth_point"] = ai_res.get("growth_point", "")
            student_data["next_goal"] = ai_res.get("next_goal", "")

            file_name, file_path = create_excel_report(student_data)
            
            if file_name and file_path:
                file_url = f"{SERVER_BASE_URL}/files/{file_name}"
                reply_text = f"【{student_data.get('name', name_query)}さんの成長レポートを作成しました】\n\n" \
                             f"■今月の成長ポイント\n{student_data.get('growth_point', '')}\n\n" \
                             f"■来月の目標\n{student_data.get('next_goal', '')}\n\n" \
                             f"📥 完成したExcelファイルのダウンロード:\n{file_url}"
            else:
                reply_text = "レポートExcelの作成に失敗しました。"

    except Exception as e:
        reply_text = f"システム処理エラーが発生しました:\n{str(e)}"

    try:
        with ApiClient(configuration) as api_client:
            line_bot_api = MessagingApi(api_client)
            line_bot_api.reply_message(
                ReplyMessageRequest(
                    reply_token=reply_token,
                    messages=[TextMessage(text=reply_text)]
                )
            )
    except Exception as e:
        print("Failed to reply to LINE:", e)

@handler.add(MessageEvent, message=TextMessageContent)
def handle_message(event):
    user_text = event.message.text
    reply_token = event.reply_token

    thread = threading.Thread(target=async_process_and_reply, args=(reply_token, user_text))
    thread.start()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
